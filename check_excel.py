#!/usr/bin/env python3
"""Excel dosyalarını kontrol et"""

import openpyxl
from pathlib import Path

def check_excel(file_path):
    """Excel dosyasını kontrol et"""
    print(f"\n{'='*60}")
    print(f"Dosya: {file_path.name}")
    print('='*60)

    wb = openpyxl.load_workbook(file_path, read_only=True)

    print(f"Toplam sheet sayisi: {len(wb.sheetnames)}")
    print(f"Ilk 5 sheet: {', '.join(wb.sheetnames[:5])}")
    print(f"Son 5 sheet: {', '.join(wb.sheetnames[-5:])}")

    # İlk sheet'i incele
    ws = wb[wb.sheetnames[0]]
    print(f"\nIlk sheet ({wb.sheetnames[0]}) boyutu:")
    print(f"  {ws.max_row} satir x {ws.max_column} sutun")

    print(f"\nIlk 3 satir:")
    for idx, row in enumerate(ws.iter_rows(max_row=3, values_only=True), 1):
        print(f"  {idx}. {row}")

    wb.close()

# Dosyaları kontrol et
base_dir = Path(__file__).parent
pdf_dir = base_dir / "Tedarikçi Dosyaları"

files = [
    "luxottica-gunes-CONVERTED.xlsx",
    "luxottica-optik-CONVERTED.xlsx"
]

for filename in files:
    file_path = pdf_dir / filename
    if file_path.exists():
        check_excel(file_path)
    else:
        print(f"\n[ERROR] Dosya bulunamadi: {filename}")

print("\n" + "="*60)
print("Kontrol tamamlandi!")
print("="*60)
